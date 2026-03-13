"""
Service de génération de fichiers Excel (XLSX) professionnels
Institut Gabriel Rita - Service Rapports
"""
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart


# === Styles de base ===
BLUE_FILL   = PatternFill('solid', fgColor='0B4B83')
GOLD_FILL   = PatternFill('solid', fgColor='E6A647')
GREEN_FILL  = PatternFill('solid', fgColor='059669')
RED_FILL    = PatternFill('solid', fgColor='DC2626')
LIGHT_FILL  = PatternFill('solid', fgColor='EFF6FF')
ALT_FILL    = PatternFill('solid', fgColor='F9FAFB')
WHITE_FILL  = PatternFill('solid', fgColor='FFFFFF')

HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
TITLE_FONT  = Font(name='Calibri', bold=True, color='0B4B83', size=14)
LABEL_FONT  = Font(name='Calibri', bold=True, color='0B4B83', size=10)
BODY_FONT   = Font(name='Calibri', size=9)
AMOUNT_FONT = Font(name='Calibri', bold=True, size=9)

THIN = Side(style='thin', color='D1D5DB')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center')
RIGHT  = Alignment(horizontal='right',  vertical='center')

AMOUNT_FMT = '#,##0 "FCFA"'


def _add_igr_header(ws, title: str, subtitle: str = '', start_row: int = 1, nb_cols: int = 6) -> int:
    """Ajouter un en-tête professionnel IGR et retourner la prochaine ligne disponible"""
    # Ligne 1 : Nom de l'école
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=nb_cols)
    cell = ws.cell(row=start_row, column=1, value='INSTITUT GABRIEL RITA — ISGR')
    cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=14)
    cell.fill = BLUE_FILL
    cell.alignment = CENTER
    ws.row_dimensions[start_row].height = 30

    # Ligne 2 : Titre du rapport
    ws.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=nb_cols)
    cell2 = ws.cell(row=start_row+1, column=1, value=title.upper())
    cell2.font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    cell2.fill = PatternFill('solid', fgColor='1B527E')
    cell2.alignment = CENTER
    ws.row_dimensions[start_row+1].height = 25

    # Ligne 3 : Sous-titre + date
    if subtitle:
        ws.merge_cells(start_row=start_row+2, start_column=1, end_row=start_row+2, end_column=nb_cols-2)
        cell3 = ws.cell(row=start_row+2, column=1, value=subtitle)
        cell3.font = Font(name='Calibri', bold=True, color='0B4B83', size=10)
        cell3.fill = PatternFill('solid', fgColor='DBEAFE')
        cell3.alignment = LEFT
    ws.merge_cells(start_row=start_row+2, start_column=nb_cols-1, end_row=start_row+2, end_column=nb_cols)
    date_cell = ws.cell(row=start_row+2, column=nb_cols-1,
                        value=f'Généré le {datetime.now().strftime("%d/%m/%Y")}')
    date_cell.font = Font(name='Calibri', italic=True, color='6B7280', size=9)
    date_cell.fill = PatternFill('solid', fgColor='DBEAFE')
    date_cell.alignment = RIGHT
    ws.row_dimensions[start_row+2].height = 20

    # Ligne vide de séparation
    ws.row_dimensions[start_row+3].height = 8

    return start_row + 4  # Prochain index de ligne


def _write_table_header(ws, row: int, headers: list, fill=None) -> None:
    """Écrire une ligne d'en-tête de tableau"""
    fill = fill or BLUE_FILL
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22


def _auto_col_widths(ws, min_w=8, max_w=40):
    """Ajuster automatiquement les largeurs de colonnes"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 4))


def _add_footer(ws, row: int, nb_cols: int):
    """Ajouter un pied de page"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nb_cols)
    footer = ws.cell(row=row, column=1,
                     value=f'Institut Gabriel Rita © {datetime.now().year} — Document confidentiel')
    footer.font = Font(name='Calibri', italic=True, color='9CA3AF', size=8)
    footer.alignment = CENTER


class ExcelService:

    @staticmethod
    def generate_schedule(data: dict) -> bytes:
        """Génère l'emploi du temps en XLSX"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"EDT {data['class_name']}"

        days = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
        time_slots = data.get('time_slots', [])
        schedule = data.get('schedule', {})
        nb_cols = 7  # Horaire + 6 jours

        # Page setup (paysage)
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 9  # A4

        # En-tête
        next_row = _add_igr_header(ws, 'Emploi du Temps',
                                    f"Classe : {data['class_name']} — {data['school_year']}",
                                    nb_cols=nb_cols)

        # En-têtes colonnes
        _write_table_header(ws, next_row, ['Horaire'] + days,
                            fill=PatternFill('solid', fgColor='1B527E'))
        next_row += 1

        # Données
        for i, ts in enumerate(time_slots):
            row_fill = ALT_FILL if i % 2 == 0 else WHITE_FILL

            # Heure
            time_cell = ws.cell(row=next_row, column=1, value=ts)
            time_cell.font = Font(name='Calibri', bold=True, color='0B4B83', size=9)
            time_cell.fill = PatternFill('solid', fgColor='DBEAFE')
            time_cell.alignment = CENTER
            time_cell.border = THIN_BORDER

            for j, day in enumerate(days):
                cell = ws.cell(row=next_row, column=j + 2)
                slot = schedule.get(day, {}).get(ts)
                if slot:
                    val = f"{slot.get('subject', '')}\n{slot.get('teacher_name', '')}"
                    cell.value = val
                    cell.font = Font(name='Calibri', bold=True, color='1E40AF', size=9)
                    cell.fill = LIGHT_FILL
                else:
                    cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = THIN_BORDER

            ws.row_dimensions[next_row].height = 40
            next_row += 1

        _add_footer(ws, next_row + 1, nb_cols)
        _auto_col_widths(ws, min_w=12)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def generate_student_report(data: dict) -> bytes:
        """Génère le relevé de compte étudiant en XLSX"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Relevé de compte'

        student = data['student']
        nb_cols = 4
        next_row = _add_igr_header(
            ws, 'Relevé de Compte Étudiant',
            f"{student['firstName']} {student['lastName']} — Matricule: {student['matricule']}",
            nb_cols=nb_cols
        )

        # Fiche étudiant
        fields = [
            ('Nom complet', f"{student['firstName']} {student['lastName']}"),
            ('Matricule', student['matricule']),
            ('Classe', student['class_name']),
            ('Frais dus', student['totalFeesDue']),
            ('Frais payés', student['totalFeesPaid']),
            ('Solde', student['balance']),
        ]
        for label, val in fields:
            ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=2)
            lbl = ws.cell(row=next_row, column=1, value=label)
            lbl.font = LABEL_FONT
            lbl.fill = LIGHT_FILL
            lbl.border = THIN_BORDER
            lbl.alignment = LEFT

            ws.merge_cells(start_row=next_row, start_column=3, end_row=next_row, end_column=4)
            val_cell = ws.cell(row=next_row, column=3, value=val)
            val_cell.border = THIN_BORDER
            val_cell.alignment = LEFT
            if isinstance(val, (int, float)):
                val_cell.number_format = AMOUNT_FMT
                val_cell.font = Font(name='Calibri', bold=True, size=10,
                                     color='059669' if val >= 0 else 'DC2626')
            next_row += 1

        # Paiements
        next_row += 1
        payments = data.get('payments', [])
        header_p = ws.cell(row=next_row, column=1, value=f'HISTORIQUE DES PAIEMENTS ({len(payments)})')
        header_p.font = TITLE_FONT
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=nb_cols)
        next_row += 1

        if payments:
            _write_table_header(ws, next_row, ['Date', 'Référence', 'Mode', 'Montant (FCFA)'])
            next_row += 1
            for i, pay in enumerate(payments):
                date_str = pay['paymentDate'].strftime('%d/%m/%Y') if hasattr(pay['paymentDate'], 'strftime') else str(pay['paymentDate'])[:10]
                row_data = [date_str, pay.get('reference', ''), pay.get('method', ''), pay['amount']]
                fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=next_row, column=col, value=val)
                    c.border = THIN_BORDER
                    c.fill = fill
                    c.font = BODY_FONT
                    if col == 4:
                        c.number_format = AMOUNT_FMT
                        c.alignment = RIGHT
                        c.font = AMOUNT_FONT
                    else:
                        c.alignment = LEFT
                next_row += 1

        _add_footer(ws, next_row + 1, nb_cols)
        _auto_col_widths(ws)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def generate_global_school(data: dict) -> bytes:
        """Génère le rapport global de l'école en XLSX"""
        wb = openpyxl.Workbook()

        # Onglet 1: Dashboard
        ws_dash = wb.active
        ws_dash.title = 'Tableau de bord'
        stats = data['statistics']
        nb_cols = 6

        next_row = _add_igr_header(ws_dash, "Rapport Global de l'École",
                                    f"Année scolaire {data['school_info']['year']}",
                                    nb_cols=nb_cols)
        kpis = [
            ('Élèves inscrits', stats['total_students']),
            ('Élèves actifs', stats['active_students']),
            ('Personnel', stats['total_staff']),
            ('Classes', stats['total_classes']),
            ('Recettes (FCFA)', stats['total_revenue']),
        ]
        fills_kpi = ['0B4B83', '059669', '7C3AED', 'D97706', '0B4B83']
        for i, ((label, val), color) in enumerate(zip(kpis, fills_kpi)):
            col = i + 1
            label_c = ws_dash.cell(row=next_row, column=col, value=label)
            label_c.font = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
            label_c.fill = PatternFill('solid', fgColor=color)
            label_c.alignment = CENTER
            label_c.border = THIN_BORDER

            val_c = ws_dash.cell(row=next_row+1, column=col, value=val)
            val_c.font = Font(name='Calibri', bold=True, color=color, size=16)
            val_c.alignment = CENTER
            val_c.border = THIN_BORDER
            if label == 'Recettes (FCFA)':
                val_c.number_format = AMOUNT_FMT
            ws_dash.row_dimensions[next_row].height = 20
            ws_dash.row_dimensions[next_row+1].height = 30
        next_row += 3

        _add_footer(ws_dash, next_row, nb_cols)
        _auto_col_widths(ws_dash)

        # Onglet 2: Étudiants
        ws_students = wb.create_sheet('Étudiants')
        next_row2 = _add_igr_header(ws_students, "Liste des Étudiants",
                                     f"{len(data['students'])} élèves", nb_cols=5)
        _write_table_header(ws_students, next_row2, ['#', 'Nom', 'Prénom', 'Classe', 'Solde (FCFA)'])
        next_row2 += 1
        for i, s in enumerate(data['students']):
            fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
            row_vals = [i+1, s['lastName'], s['firstName'], s['className'], s['balance']]
            for col, val in enumerate(row_vals, 1):
                c = ws_students.cell(row=next_row2, column=col, value=val)
                c.border = THIN_BORDER
                c.fill = fill
                c.font = BODY_FONT
                c.alignment = LEFT
                if col == 5:
                    c.number_format = AMOUNT_FMT
                    c.font = Font(name='Calibri', bold=True, size=9,
                                  color='059669' if s['balance'] >= 0 else 'DC2626')
            next_row2 += 1
        _add_footer(ws_students, next_row2+1, 5)
        _auto_col_widths(ws_students)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def generate_late_payments(data: dict) -> bytes:
        """Génère le rapport des paiements en retard en XLSX"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Paiements en retard'

        late = data.get('late_payments', [])
        nb_cols = 5
        total = sum(p['balance'] for p in late)

        next_row = _add_igr_header(ws, 'Rapport — Paiements en Retard',
                                    f"{len(late)} élève(s) — Total impayé: {total:,.0f} FCFA",
                                    nb_cols=nb_cols)

        # Résumé
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=nb_cols)
        sum_cell = ws.cell(row=next_row, column=1,
                           value=f'TOTAL IMPAYÉ : {total:,.0f} FCFA — {len(late)} élève(s) concerné(s)')
        sum_cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        sum_cell.fill = RED_FILL
        sum_cell.alignment = CENTER
        sum_cell.border = THIN_BORDER
        ws.row_dimensions[next_row].height = 25
        next_row += 2

        _write_table_header(ws, next_row, ['#', 'Nom', 'Prénom', 'Classe', 'Montant dû (FCFA)'])
        next_row += 1
        for i, s in enumerate(late):
            fill = PatternFill('solid', fgColor='FEF2F2') if i % 2 == 0 else WHITE_FILL
            row_vals = [i+1, s['lastName'], s['firstName'], s['className'], s['balance']]
            for col, val in enumerate(row_vals, 1):
                c = ws.cell(row=next_row, column=col, value=val)
                c.border = THIN_BORDER
                c.fill = fill
                c.font = BODY_FONT
                c.alignment = LEFT
                if col == 5:
                    c.number_format = AMOUNT_FMT
                    c.font = Font(name='Calibri', bold=True, color='DC2626', size=9)
                    c.alignment = RIGHT
            next_row += 1

        # Total
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
        ws.cell(row=next_row, column=1, value='TOTAL').font = Font(name='Calibri', bold=True, size=10)
        tot_c = ws.cell(row=next_row, column=5, value=total)
        tot_c.number_format = AMOUNT_FMT
        tot_c.font = Font(name='Calibri', bold=True, color='DC2626', size=10)
        tot_c.border = Border(top=Side(style='medium', color='DC2626'))

        _add_footer(ws, next_row+2, nb_cols)
        _auto_col_widths(ws)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def generate_moratoriums(data: dict) -> bytes:
        """Génère le rapport des moratoires en XLSX"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Moratoires'

        students = data.get('students', [])
        nb_cols = 5
        next_row = _add_igr_header(ws, 'Rapport — Moratoires',
                                    f"{len(students)} moratoire(s) — Année {data['school_year']}",
                                    nb_cols=nb_cols)

        if students:
            _write_table_header(ws, next_row, ['#', 'Nom', 'Prénom', 'Classe', 'Solde (FCFA)'],
                                fill=PatternFill('solid', fgColor='D97706'))
            next_row += 1
            for i, s in enumerate(students):
                fill = PatternFill('solid', fgColor='FFF7ED') if i % 2 == 0 else WHITE_FILL
                row_vals = [i+1, s['lastName'], s['firstName'], s['className'], s['balance']]
                for col, val in enumerate(row_vals, 1):
                    c = ws.cell(row=next_row, column=col, value=val)
                    c.border = THIN_BORDER
                    c.fill = fill
                    c.font = BODY_FONT
                    c.alignment = LEFT
                    if col == 5:
                        c.number_format = AMOUNT_FMT
                        c.font = Font(name='Calibri', bold=True, color='DC2626', size=9)
                next_row += 1

        _add_footer(ws, next_row+1, nb_cols)
        _auto_col_widths(ws)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def generate_payments_by_class(data: dict) -> bytes:
        """Génère le rapport des paiements par classe en XLSX"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Classe {data['class_name']}"

        nb_cols = 6
        next_row = _add_igr_header(ws, f"Paiements — Classe {data['class_name']}",
                                    f"Année {data['school_year']}",
                                    nb_cols=nb_cols)

        # Résumé
        summaries = [
            ('Total facturé', data['total_invoiced'], '0B4B83'),
            ('Total payé', data['total_paid'], '059669'),
            ('Reste à payer', data['total_invoiced'] - data['total_paid'], 'DC2626'),
        ]
        for i, (label, val, color) in enumerate(summaries):
            col = i * 2 + 1
            ws.merge_cells(start_row=next_row, start_column=col, end_row=next_row, end_column=col+1)
            lbl = ws.cell(row=next_row, column=col, value=label)
            lbl.font = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
            lbl.fill = PatternFill('solid', fgColor=color)
            lbl.alignment = CENTER
            lbl.border = THIN_BORDER

            ws.merge_cells(start_row=next_row+1, start_column=col, end_row=next_row+1, end_column=col+1)
            val_c = ws.cell(row=next_row+1, column=col, value=val)
            val_c.number_format = AMOUNT_FMT
            val_c.font = Font(name='Calibri', bold=True, color=color, size=13)
            val_c.alignment = CENTER
            val_c.border = THIN_BORDER
        next_row += 3

        students = data.get('students', [])
        _write_table_header(ws, next_row, ['Nom', 'Prénom', 'Matricule', 'Nb Paiements', 'Solde (FCFA)', 'Statut'])
        next_row += 1
        for i, s in enumerate(students):
            fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
            statut = '✅ Payé' if s['balance'] <= 0 else '⚠️ En retard'
            row_vals = [s['lastName'], s['firstName'], s['matricule'], len(s.get('payments', [])), s['balance'], statut]
            for col, val in enumerate(row_vals, 1):
                c = ws.cell(row=next_row, column=col, value=val)
                c.border = THIN_BORDER
                c.fill = fill
                c.font = BODY_FONT
                c.alignment = LEFT
                if col == 5:
                    c.number_format = AMOUNT_FMT
                    c.font = Font(name='Calibri', bold=True, size=9,
                                  color='059669' if s['balance'] <= 0 else 'DC2626')
            next_row += 1

        _add_footer(ws, next_row+1, nb_cols)
        _auto_col_widths(ws)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
